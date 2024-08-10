import os
import sqlite3
import openpyxl


def export_to_sqlite():
    '''Экспорт данных из xlsx в sqlite'''

    prj_dir = os.path.abspath(os.path.curdir)

    a = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    base_name = 'db.sqlite3'

    connect = sqlite3.connect(prj_dir + '/' + base_name)
    cursor = connect.cursor()

    cursor.execute('CREATE TABLE IF NOT EXISTS campaigns (id int, hashtag text, goal int, collected int , user_count int, status int, charity_id int, help_receiver_count int, link_open_event_count int, published_at datatime, finished_at datatime, finish_payment_id int, title text, created datatime, updated datatime)')

    file_to_read = openpyxl.load_workbook('datasets/campaigns.xlsx', data_only=True)
    sheet = file_to_read['Query result']


    for row in range(2, sheet.max_row + 1):
        
        data = []
        
        for col in range(1, 16):
            
            value = sheet.cell(row, col).value
            
            data.append(value)

        cursor.execute("INSERT INTO campaigns VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);", (data[0], data[1], data[2], data[3], data[4], data[5], data[6], data[7], data[8], data[9], data[10], data[11], data[12], data[9], data[10]))

    connect.commit()
    connect.close()


def clear_base():
    '''Очистка базы sqlite'''

    # Получаем текущую папку проекта
    prj_dir = os.path.abspath(os.path.curdir)

    # Имя базы
    base_name = 'auto.sqlite3'

    connect = sqlite3.connect(prj_dir + '/' + base_name)
    cursor = connect.cursor()

    # Запись в базу, сохранение и закрытие соединения
    cursor.execute("DELETE FROM campaigns")
    connect.commit()
    connect.close()


export_to_sqlite()
